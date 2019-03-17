import numpy as np
import pandas as pd
import os
import glob
import ast
import csv
import time
import interruptionFinder
import openpyxl as openpyxl

'''
    Reads a map from a given input file
'''
def read_map(input_file):
    with open(input_file, 'rt') as infile:
        grid1 = np.array([list(line.strip()) for line in infile.readlines()])
    print('Grid shape', grid1.shape)

    grid1[grid1 == '@'] = 1  # object on the map
    grid1[grid1 == 'T'] = 1  # object on the map
    grid1[grid1 == '.'] = 0  # free on map

    return np.array(grid1.astype(np.int))


'''
    Read the agents original plan, i.e., their routes
'''
def read_routes_from_file(input_file):
    main_routes = []

    # reading csv file
    with open(input_file, 'rt') as csvfile:
        # creating a csv reader object
        csvreader = csv.reader(csvfile)
        # extracting each data row one by one
        for row in csvreader:
            tmp_route = []
            for y in range(0, len(row)):
                if row[y]:
                    tmp_route.append(ast.literal_eval(row[y]))
            main_routes.append(tmp_route)
    return main_routes


'''
Run a single experiment
'''


def main():
    # Experiment parameters
    k=3
    TIMEOUT = 180 # seconds

    # Read the map
    map = read_map('room1.txt')

    # For every routes file, start an experiment
    path = 'Data\\'
    extension = 'csv'
    os.chdir(path)
    files = [i for i in glob.glob('*.{}'.format(extension))]

    writer = pd.ExcelWriter('Results.xlsx', engine='openpyxl')
    counter = 0

    # Call a Workbook() function of openpyxl
    # to create a new blank Workbook object
    wb = openpyxl.Workbook()

    # Get workbook active sheet
    # from the active attribute.
    sheet = wb.active

    for file in files:

        original_routes = read_routes_from_file(file)
        org_total_g=0
        for x in range(0, len(original_routes)):
            org_total_g = org_total_g + len(original_routes[x])-1

        for k in range(1, 7):
            print('********', counter, '********')
            print(file)
            new_total_g = 0
            counter = counter + 1

            solution = []

            start_time = time.time()

            solver = interruptionFinder.InterruptionFinder()
            solution, route, total_expanded_nodes, runtime = solver.search_for_interrupt_plan(map,k,original_routes, TIMEOUT)

            elapsed_time = time.time() - start_time
            if elapsed_time<TIMEOUT:
                print('Len of the route:', len(route[1]))
                print('No. of agents:', len(route))
                print('No. of errors:', k)
                print('Total runtime: ' + str(elapsed_time))
                total_routes_cost = sum([len(route) - 1 for route in solution[0].route])
                # print('Total routes cumulative sum: '+ str(total_routes_cost))
                print('Total expanded nodes: ' + str(total_expanded_nodes))
                print('Number of interruption for best solution: ' + str(k - solution[-1].k))

                for x in range(0, len(route)):
                    # print('**** Agent ', x, '****')
                    start = route[x][0][0]
                    goal = route[x][-1][0]
                    first_time_goal = 1
                    colors = ['r', 'purple', 'b', 'c', 'm', 'y', 'k', 'w']

                    # extract x and y coordinates from route list
                    x_coords = []
                    y_coords = []
                    for y in range(0, len(solution)):

                        x1 = solution[y].pos[x][0]
                        y1 = solution[y].pos[x][1]

                        if x1 == (route[x][-1][0][0]) and y1 == (route[x][-1][0][1]):
                            if first_time_goal == 0:
                                continue
                            else:
                                first_time_goal = 0
                        else:
                            new_total_g = new_total_g + 1

                        x_coords.append(x1)
                        y_coords.append(y1)

                print('Original g_score is - ', org_total_g)
                print('New g_score is - ', new_total_g)

                no_of_interrupt = k - solution[-1].k

                res = [file, k, len(route), len(route[1]), elapsed_time, org_total_g, total_expanded_nodes,
                       no_of_interrupt,
                       new_total_g]
            else: # Finished before the time limit
                print('***** Time Out *****')
                print('Len of the route:', len(route[1]))
                print('No. of agents:', len(route))
                print('No. of errors:', k)
                elapsed_time = time.time() - start_time
                print('Total runtime: ' + str(elapsed_time))

                res = [file, k, 'time out', 'time out', elapsed_time, org_total_g, total_expanded_nodes,
                       no_of_interrupt,
                       'time out']

            # writing to the specified cell
            for len_of_row in range(1, 10):
                sheet.cell(row=counter, column=len_of_row).value = res[len_of_row - 1]
            wb.save('results.xlsx')


if __name__ == "__main__":
    main()