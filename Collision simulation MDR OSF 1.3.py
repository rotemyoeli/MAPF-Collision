import math
import numpy as np
import heapq
import matplotlib.pyplot as plt
import matplotlib.ticker as plticker
import csv
import ast
import sys
import time
import copy
import warnings
import glob
import os
import glob
import pandas as pd
import openpyxl as openpyxl
warnings.filterwarnings("ignore")



# Call a Workbook() function of openpyxl
# to create a new blank Workbook object
wb = openpyxl.Workbook()

# Get workbook active sheet
# from the active attribute.
sheet = wb.active



# load the map file into numpy array
with open('Room1.txt', 'rt') as infile:
    grid1 = np.array([list(line.strip()) for line in infile.readlines()])
print('Grid shape', grid1.shape)

grid1[grid1 == '@'] = 1  # object on the map
grid1[grid1 == 'T'] = 1  # object on the map
grid1[grid1 == '.'] = 0  # free on map

grid = np.array(grid1.astype(np.int))

# Plot the map
fig, ax = plt.subplots(figsize=(10, 5))
ax.imshow(grid, cmap=plt.cm.ocean)

# To convert new xlsx files into csv format, to use only once
#path = 'C:\\Users\\rotem\\Dropbox\\Studies\\Thesis\\Code\\Ver 2.0\\data\\'
#extension = 'xlsx'
#os.chdir(path)
#excel_files = [i for i in glob.glob('*.{}'.format(extension))]
#for excel in excel_files:
#    out = excel.split('.')[0]+'.csv'
#    df = pd.read_excel(excel) # if only the first sheet is needed.
#    df.to_csv(out)

    # If you know the name of the column skip this
#    first_column = df.columns[0]
#    # Delete first
#    df = df.drop([first_column], axis=1)
#    df.to_csv(out, index=False)


path = 'C:\\Users\\rotem\\Dropbox\\Studies\\Thesis\\Code\\Ver 2.0\\data\\'
extension = 'csv'
os.chdir(path)
files = [i for i in glob.glob('*.{}'.format(extension))]

##############################################################################
# heuristic function for A* path scoring
##############################################################################

def heuristic(a, b):
    return (b[0] - a[0]) ** 2 + (b[1] - a[1]) ** 2

##############################################################################
# heuristic function for heuristic_interrupt path scoring
##############################################################################


'''
search_node contains position of all agent, allowed abnormal moves, current time step
routes contains the planned routes for all agents [ including the path they passed so far ]
'''
def f_interrupt(search_node, routes):

    cost_without_interrupts = sum([len(route) for route in routes])

    # An overestimate of the cost added by the abnormal moves
    remaining_abnormal_moves = search_node.k
    num_of_agents = len(search_node.pos)

    # The idea is that the maximal addition for each abnormal move is that it will delay all agents by one time step.
    return cost_without_interrupts + num_of_agents * remaining_abnormal_moves



##############################################################################
# path finding function
##############################################################################

def astar(array, start, goal, y, route, step):
    neighbors = [(0, 1), (0, -1), (1, 0), (-1, 0), (1, 1), (1, -1), (-1, 1), (-1, -1), (0, 0)]
    #neighbors = [(0, 1), (0, -1), (1, 0), (-1, 0)]
    start_node = (start)

    close_set = set()
    came_from = {start_node: 0}
    gscore = {start_node: 0}
    fscore = {start_node: heuristic(start[0], goal)}
    oheap = []
    heapq.heappush(oheap, (fscore[start_node], start_node))
    is_collision = 0

    while oheap:

        current = heapq.heappop(oheap)[1]

        if current[0] == goal:
            data = []
            while current in came_from:
                data.append(current)
                current = came_from[current]
            data = list(reversed(data))
            return data

        close_set.add(current)

        # Add stay at current position option
        neighbor = ((current[0][0], current[0][1]), current[1] + 1)
        came_from[neighbor] = (current)
        gscore[neighbor] = gscore[current] + 1
        fscore[neighbor] = gscore[current] + 1 + heuristic(neighbor[0], goal)
        heapq.heappush(oheap, (fscore[neighbor], neighbor))

        for i, j in neighbors:

            neighbor = ((current[0][0] + i, current[0][1] + j), current[1] + 1)

            tentative_g_score = gscore[current] + heuristic(current[0], neighbor[0])

            ##Check for collision
            for a in range(0, y): #[k for k in range(0, len(route)) if k != y]:
                if len(route[a]) > neighbor[1]:
                    if neighbor[0] == route[a][neighbor[1]][0]:
                        is_collision = 1
                        break
            if is_collision == 1:
                is_collision = 0
                continue

            if 0 <= neighbor[0][0] < array.shape[0]:
                if 0 <= neighbor[0][1] < array.shape[1]:
                    if array[neighbor[0][0]][neighbor[0][1]] == 1:
                        # array bound 1 on map
                        continue
                else:
                    # array bound y walls
                    continue
            else:
                # array bound x walls
                continue

            if neighbor in close_set and tentative_g_score >= gscore.get(neighbor[0], 0):
                continue

            if tentative_g_score < gscore.get(neighbor[0], 0) or neighbor not in [i[1] for i in oheap]:
                came_from[neighbor] = (current)
                gscore[neighbor] = tentative_g_score
                fscore[neighbor] = tentative_g_score + heuristic(neighbor[0], goal)
                heapq.heappush(oheap, (fscore[neighbor], neighbor))

    return False

class Node:

    def __init__(self, pos, k, step, parent, route=None):
        self.pos = pos
        self.k = k
        self.step = step
        self.parent = parent
        self.route = route

    def weight(self):

        next_step_radians = math.atan2((self.route[0][-1][0][1]) - self.pos[0][1],
                                       (self.route[0][-1][0][0]) - self.pos[0][0])
        w = math.degrees(next_step_radians)

        return w

    def __lt__(self, other):
        return self.weight() < other.weight()

'''
 route = list containing the planned route for each agent 
'''
def search_for_interrupt_plan(route):
    total_expanded_nodes = 0
    actions = [(0, 1), (0, -1), (1, 0), (-1, 0), (1, 1), (1, -1), (-1, 1), (-1, -1), (0, 0)]
    #actions = [(0, 1), (0, -1), (1, 0), (-1, 0)]
    open = []
    pos = list()
    step = 0

    # Set starting positions for all agents into a node
    for x in range(0, len(route)):
        pos.append(route[x][0][0])

    # pos = a list that contains location of all agents
    # k = number of allowed abnormal moves
    # step = current time step
    # None = the parent pointer

    start = Node(pos=pos, k=k, step=step, parent=None,route=route)

    f_score = f_interrupt(start, route)
    fscore = {start: f_score}
    heapq.heappush(open, (fscore[start], start))

    goals = list()

    best_goal_value = 0
    for v in range(0, len(route)):
        best_goal_value = best_goal_value + len(route[v])

    while (len(open)>0):
        node = heapq.heappop(open)[1]

        #count total expanded nodes
        total_expanded_nodes = total_expanded_nodes + 1
        step = node.step + 1

        # No need to expand a node that cannot improve the best goal
        if f_interrupt(node, node.route) <= best_goal_value:
            continue
        if is_goal(node, node.route):
            continue

        #print("%d" % (len(goals)))

        for action in actions:

            #TODO: change to be from current to next planned step
            if step < len(route[0]):
                planned_radians = math.atan2(route[0][step][0][1] - node.pos[0][1], route[0][step][0][0] - node.pos[0][0])
                planned_degrees = math.degrees(planned_radians)
            else:
                planned_radians = math.atan2(route[0][-1][0][1] - node.pos[0][1], route[0][-1][0][0] - node.pos[0][0])
                planned_degrees = math.degrees(planned_radians)
            #print(planned_degrees)

            next_step_radians = math.atan2((node.pos[0][1] + action[1]) - node.pos[0][1],
                                           (node.pos[0][0] + action[0]) - node.pos[0][0])
            next_step_degrees = math.degrees(next_step_radians)

            anglediff = (next_step_degrees - planned_degrees + 180 + 360) % 360 - 180
            #print(anglediff, next_step_degrees, planned_degrees)

            if (anglediff > 90 or anglediff < -90):
                continue

            if node.k == 0:
                #no more err to use
                pos = list()
                for x in range(0, len(node.route)):
                    if step < len(node.route[x]):
                        pos.append(node.route[x][step][0])
                    else:
                        #TODO: Try to add none instead the last node and verify the code
                        pos.append(node.route[x][-1][0])
                new_node = Node(pos, 0, step, node,node.route)
                if (is_goal(new_node, new_node.route)):
                    goals.append(new_node)

                    # find the total cost of the current route
                    cost_new_node = 0
                    for m in range(0, len(new_node.route)):
                        cost_new_node = cost_new_node + new_node.route[m][-1][1]
                    if cost_new_node > best_goal_value:
                        best_goal_value = cost_new_node
                else:
                    f_score = f_interrupt(new_node, route)
                    fscore[new_node] = f_score + node.step# + (np.random.random()/1)
                    heapq.heappush(open, (fscore[new_node], new_node))

                break

            # if action is legal (not collide in wall)
            if grid[node.pos[0][0] + action[0]][node.pos[0][1] + action[1]] == 1:
                # array bound 1 on map
                continue

            new_node, _ = apply(action, node, step, node.route)
            if new_node == None:
                continue
            if (is_goal(new_node, new_node.route)):
                goals.append(new_node)

                # find the total cost of the current route
                cost_new_node = 0
                for m in range(0, len(new_node.route)):
                    cost_new_node = cost_new_node + new_node.route[m][-1][1]
                if cost_new_node > best_goal_value:
                    best_goal_value = cost_new_node
            else:
                f_score = f_interrupt(new_node, new_node.route)
                fscore[new_node] = f_score + node.step# + (np.random.random()/1)
                heapq.heappush(open, (fscore[new_node], new_node))

    # After all goals have been found
    max_step = 0
    best_goal = None
    for goal in goals:
        if goal.step > max_step:
            max_step = goal.step
            best_goal = goal

    # Reconstruct the solution
    solution = [best_goal]
    parent = best_goal.parent

    while parent is not None:
        solution.append(parent)
        parent = parent.parent

    #solution = list(reversed(solution))

    return (solution, route,total_expanded_nodes)


def is_goal(node, route):

    for x in range(0, len(node.pos)):
        if node.pos[x] is not route[x][-1][0] and node.pos[x] is not None:
            return False
    return True

def apply(action, node, step, route_):
    route = copy.deepcopy(route_)
    tmp_node = list()
    pos = list()

    if step < len(route[0]):
        #If action of Agent 0 is according the plan - don't decrease the no. of remained error
        if route[0][step][0] == (node.pos[0][0] + action[0], node.pos[0][1] + action[1]):
            pos.append(route[0][step][0])
            for w in range(1, len(route)):
                if step < len(route[w]):
                    pos.append(route[w][step][0])
                else:
                    pos.append(route[w][-1][0])
            tmp_node = Node(pos, node.k, step, node, route=route)
            return tmp_node, route

        #If action of Agent 0 is not according the plan - decrease the no. of remained error and sent to re-calculate new routes
        else:
            tmp_node, route = calc_new_routes(action, node, step, route)

    else:
        return None, None

    return tmp_node, route

def calc_new_routes(action, node, step, route_):
    route = copy.deepcopy(route_)
    tmp_node = list()
    pos = list()
    # set the no. of agent that making the move for interruptions
    a = 0


    tmp_node.append((node.pos[0][0] + action[0], node.pos[0][1] + action[1]))


    ###########################################################################
    # Send Agent - 0 to A* (grid, start = next new step, goal, agent no., step)
    ###########################################################################

    # If there are more steps, (i.e. it is not at the end of it's route) for agent 0,
    # so run A* for Agent - 0
    if step < len(route[0]):
        new_route = []
        for h in range(0, step):
            new_route.append(route[a][h])

        start = (((node.pos[0][0] + action[0], node.pos[0][1] + action[1])), step)
        goal = route[a][-1][0]

        if step >= 1:
            route[a] = astar(grid, start, goal, a, route, step)
            for h in route[a]:
                new_node = ((h[0]), new_route[-1][1] + 1)
                new_route.append(new_node)
            route[a] = new_route
        else:
            route[a] = astar(grid, start, goal, a, route, step)
    else:
        new_node = route[a][-1][0]
        route[a].append(new_node)

    ##########################################################
    # Build the new Node after agent - 0 build its new route.
    ##########################################################

    pos.append(tmp_node[0])
    for x in range(1, len(route)):
        if step < len(route[x]):
            pos.append(route[x][step][0])
        else:
            pos.append(route[x][-1][0])
    tmp_node = Node(pos, node.k-1, step, node,route=route)

    ##########################################################
    # After agent - 0 re-calc its new route then check if there are collisions
    # between agent - 0 and any other. if there is, then send agent x to A* and check again.
    # After that, set positions for all agents into a node
    ##########################################################

    for x in range(0, len(tmp_node.pos)):
        for y in range(x + 1, len(tmp_node.pos)):

            if tmp_node.pos[x] == tmp_node.pos[y] and x != y:

                # if more steps for agent y
                if step < len(route[y]):
                    new_route = []
                    for h in range(0, step):
                        new_route.append(route[y][h])

                    start = route[y][step - 1]
                    goal = route[y][-1][0]

                    if step >= 1:
                        route[y] = astar(grid, start, goal, y, route, step)
                        del route[y][0]
                        for h in route[y]:
                            new_node = ((h[0]), new_route[-1][1] + 1)
                            new_route.append(new_node)
                        route[y] = new_route
                    else:
                        route[y] = astar(grid, start, goal, y, route, step)
                else:
                    new_node = (route[y][-1][0], step)
                    route[y].append(new_node)

                tmp_node = list()
                pos = list()
                tmp_node.append((node.pos[0][0] + action[0], node.pos[0][1] + action[1]))
                pos.append(tmp_node[0])
                for x in range(1, len(route)):
                    if step < len(route[x]):
                        pos.append(route[x][step][0])
                    else:
                        pos.append(route[x][-1][0])
                tmp_node = Node(pos, node.k - 1, step, node, route=route)
    #
    ###########################################################

    return tmp_node, route


counter = 1

res = []
res = ['File Name', 'Leader No.', 'No. of Errors', 'No. of Agents', 'Len of Route', 'Run Time', 'Org. Cost', 'Total Expanded Nodes', 'No of Interrupt',
                   'New Score']
# writing to the specified cell
for len_of_row in range(1, 11):
    sheet.cell(row=counter, column=len_of_row).value = res[len_of_row-1]
wb.save('results.xlsx')
counter = counter+1


for file in files:
    main_route = []
    new_route = []
    tmp_route = []
    #k = 3


    total_expanded_nodes = 0
    org_total_g = 0
    new_total_g = 0


    # reading csv file
    with open(file, 'rt') as csvfile:
        # creating a csv reader object
        csvreader = csv.reader(csvfile)
        header = next(csvreader)
        # extracting each data row one by one
        for row in csvreader:
            tmp_route = []
            for y in range(0, len(row)):
                if row[y]:
                    tmp_route.append(ast.literal_eval(row[y]))
            main_route.append(tmp_route)


    for x in range(0, len(main_route)):
        org_total_g = org_total_g + len(main_route[x])


    ######LOOP Every agents as the leader######

    for current_agent in range(1, len(main_route)):
        temp_list_first = []
        temp_list_current = []
        temp_list_current = main_route[current_agent]
        temp_list_first = main_route[0]
        main_route[0] = temp_list_current
        main_route[current_agent] = temp_list_first


        ######LOOP for k between 1-5 errors######

        for k in range(1, 3):

            solution = []

            start_time = time.time()

            num_of_interrupts = 0

            solution, route, total_expanded_nodes = search_for_interrupt_plan(main_route)

            elapsed_time = time.time() - start_time

            #for l in range(0, len(route)):
            #    print(route[l])

            print(file)
            print(counter-1)

            print('Total runtime: '+ str(elapsed_time))
            total_routes_cost = sum([len(route) for route in solution[0].route])
            #print('Total routes cumulative sum: '+ str(total_routes_cost))
            print('Total expanded nodes: '+ str(total_expanded_nodes))
            print('Number of interruption for best solution: '+str(k - solution[-1].k))
            print('Number of Errors: ', k)
            new_total_g = 0


            for x in range(0, len(route)):
                #print('**** Agent ', x, '****')
                start = route[x][0][0]
                goal = route[x][-1][0]
                first_time_goal = 1
                colors = ['r', 'purple', 'b', 'c', 'm', 'y', 'k', 'w']
                ##############################################################################
                # plot the path
                ##############################################################################

                # extract x and y coordinates from route list
                x_coords = []
                y_coords = []
                for y in range(0, len(solution)):

                    x1 = solution[y].pos[x][0]
                    y1 = solution[y].pos[x][1]

                    if x1 == (route[x][-1][0][0]) and y1 == (route[x][-1][0][1]):
                        if first_time_goal == 0:
                            continue
                        else:
                            first_time_goal = 0
                    #else:
                    #    new_total_g = new_total_g + 1

                    x_coords.append(x1)
                    y_coords.append(y1)
                new_total_g = new_total_g + len(x_coords)

            print('Original g_score is - ', org_total_g)
            print('New g_score is - ', new_total_g)

            #for v in range(0, len(route)):
            #    print(route[v])


            no_of_interrupt = k - solution[-1].k
            res = [file, current_agent, k, len(route), len(route[1]), elapsed_time, org_total_g, total_expanded_nodes, no_of_interrupt,
                   new_total_g]

            # writing to the specified cell
            for len_of_row in range(1, 11):
                sheet.cell(row=counter, column=len_of_row).value = res[len_of_row - 1]
            counter = counter + 1

    wb.save('results.xlsx')


