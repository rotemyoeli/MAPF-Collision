import math
import numpy as np
import heapq
import matplotlib.pyplot as plt
import matplotlib.ticker as plticker
import csv
import ast
import sys
import time
import copy
import warnings
import glob
import os
import glob
import pandas as pd
import openpyxl as openpyxl
warnings.filterwarnings("ignore")
import utils




class MDR():


    def _compute_makespan(routes):
        return max([len(route) for route in routes])

    ##############################################################################
    # heuristic function for heuristic_interrupt path scoring
    ##############################################################################
    '''
    search_node contains position of all agent, allowed abnormal moves, current time step
    routes contains the planned routes for all agents [ including the path they passed so far ]
    '''
    def f_interrupt(self, search_node, routes):

        cost_without_interrupts = MDR._compute_makespan(routes)

        # An overestimate of the cost added by the abnormal moves
        remaining_abnormal_moves = search_node.k
        num_of_agents = len(search_node.pos)

        # The idea is that the maximal addition for each abnormal move is that it will delay all agents by one time step.
        return -1*(cost_without_interrupts + num_of_agents * remaining_abnormal_moves)  # TODO: CHANGE 2

        '''
         route = list containing the planned route for each agent 
        '''

    def is_goal(self, node, route):

        for x in range(0, len(node.pos)):
            if node.pos[x] is not route[x][-1][0] and node.pos[x] is not None:
                return False
        return True

    def apply(self, action, node, step, route_):
        #route = copy.deepcopy(route_)
        routes = [route for route in route_]
        tmp_node = list()
        pos = list()

        if step < len(routes[0]):
            # If action of Agent 0 is according the plan - don't decrease the no. of remained error
            if routes[0][step][0] == (node.pos[0][0] + action[0], node.pos[0][1] + action[1]):
                pos.append(routes[0][step][0])
                for w in range(1, len(routes)):
                    if step < len(routes[w]):
                        pos.append(routes[w][step][0])
                    else:
                        pos.append(routes[w][-1][0])
                tmp_node = MDR.Node(pos, node.k, step, node, route=routes)
                return tmp_node, routes

            # If action of Agent 0 is not according the plan - decrease the no. of remained error and sent to re-calculate new routes
            else:
                tmp_node, routes = self.calc_new_routes(action, node, step, routes)

        else:
            return None, None

        return tmp_node, routes

    def calc_new_routes(self, action, node, step, route_):
        #route = copy.deepcopy(route_)
        routes = [route for route in route_]
        tmp_node = list()
        pos = list()
        # set the no. of agent that making the move for interruptions
        a = 0

        tmp_node.append((node.pos[0][0] + action[0], node.pos[0][1] + action[1]))

        ###########################################################################
        # Send Agent - 0 to A* (grid, start = next new step, goal, agent no., step)
        ###########################################################################

        # If there are more steps, (i.e. it is not at the end of it's route) for agent 0,
        # so run A* for Agent - 0
        if step < len(routes[0]):
            new_route = []
            for h in range(0, step):
                new_route.append(routes[a][h])

            start = (((node.pos[0][0] + action[0], node.pos[0][1] + action[1])), step)
            goal = routes[a][-1][0]

            if step >= 1:
                routes[a] = astar(self.grid, start, goal, a, routes, step, self.db_heuristic)
                for h in routes[a]:
                    new_node = ((h[0]), new_route[-1][1] + 1)
                    new_route.append(new_node)
                routes[a] = new_route
            else:
                routes[a] = astar(self.grid, start, goal, a, routes, step, self.db_heuristic)
        else:
            new_node = routes[a][-1][0]
            routes[a].append(new_node)

        ##########################################################
        # Build the new Node after agent - 0 build its new route.
        ##########################################################

        pos.append(tmp_node[0])
        for x in range(1, len(routes)):
            if step < len(routes[x]):
                pos.append(routes[x][step][0])
            else:
                pos.append(routes[x][-1][0])
        tmp_node = MDR.Node(pos, node.k - 1, step, node, route=routes)

        ##########################################################
        # After agent - 0 re-calc its new route then check if there are collisions
        # between agent - 0 and any other. if there is, then send agent x to A* and check again.
        # After that, set positions for all agents into a node
        ##########################################################

        for x in range(0, len(tmp_node.pos)):
            for y in range(x + 1, len(tmp_node.pos)):

                if tmp_node.pos[x] == tmp_node.pos[y] and x != y:

                    # if more steps for agent y
                    if step < len(routes[y]):
                        new_route = []
                        for h in range(0, step):
                            new_route.append(routes[y][h])

                        start = routes[y][step - 1]
                        goal = routes[y][-1][0]

                        if step >= 1:
                            routes[y] = astar(self.grid, start, goal, y, routes, step, self.db_heuristic)
                            #print(y, step, routes[y])
                            # print(y, route)
                            del routes[y][0]

                            for h in routes[y]:
                                new_node = ((h[0]), new_route[-1][1] + 1)
                                new_route.append(new_node)
                            routes[y] = new_route
                        else:
                            routes[y] = astar(self.grid, start, goal, y, routes, step, self.db_heuristic)
                            #print(y, step, routes[y])
                    else:
                        new_route = []
                        for h in range(0, len(routes[y])):
                            new_route.append(routes[y][h])
                        new_node = (routes[y][-1][0], step)
                        new_route.append(new_node)
                        routes[y] = new_route
                        #routes[y].append(new_node) # BUG: updating the route instead of creating a new one

                    tmp_node = list()
                    pos = list()
                    tmp_node.append((node.pos[0][0] + action[0], node.pos[0][1] + action[1]))
                    pos.append(tmp_node[0])
                    for x in range(1, len(routes)):
                        if step < len(routes[x]):
                            pos.append(routes[x][step][0])
                        else:
                            pos.append(routes[x][-1][0])
                    tmp_node = MDR.Node(pos, node.k - 1, step, node, route=routes)
        #
        ###########################################################
        #count_1 = 0
        #for c in range(0, len(routes)):
            #count_1 = count_1 + 1
            #for d in range(c + 1, len(routes)):
                #print(routes[c][count_1][0], routes[d][count_1][0])
                #if routes[c][count_1][0] == routes[d][count_1][0]:
                    #print(count_1,'C :', c, routes[c])
                    #print(count_1, 'D :', d, routes[d])

        return tmp_node, routes

    def search_for_interrupt_plan(self, grid, route, k):
        total_expanded_nodes = 0
        actions = [(0, 1), (0, -1), (1, 0), (-1, 0), (0, 0)]

        open = []
        pos = list()
        step = 0
        self.grid = grid

        # Set starting positions for all agents into a node
        num_of_agents = len(route)
        for x in range(0, num_of_agents ):
            pos.append(route[x][0][0])

        print("Pre-computing shortest paths from goals...", end='')
        goal_to_db = dict()
        for x in range(0, num_of_agents):
            agent_goal = route[x][-1][0]
            operators_and_costs = [(0, 1,1), (0, -1,1), (1, 0,1), (-1, 0,1)]
            goal_to_db[agent_goal] = utils.dijkstra(self.grid, agent_goal, operators_and_costs)

        self.db_heuristic = DatabaseHeuristic(goal_to_db)
        print("Done")


        start = MDR.Node(pos=pos, k=k, step=step, parent=None, route=route)
        f_score = self.f_interrupt(start, route)
        fscore = {start: f_score}
        heapq.heappush(open, (fscore[start], start))

        goals = list()
        best_goal_value = MDR._compute_makespan(route) # TODO Change

        while (len(open) > 0):
            # print(len(open))
            (f, node) = heapq.heappop(open)

            # count total expanded nodes
            # print(node.pos,f, len(open))
            total_expanded_nodes = total_expanded_nodes + 1
            if total_expanded_nodes % 10000 ==0:
                print("Total expanded nodes so far = %d, max f in OPEN = %d" % (total_expanded_nodes, f))
            step = node.step + 1

            # No need to expand a node that cannot improve the best goal
            node_value = -f
            if node_value < best_goal_value: # TODO: Change 4
                continue
            if self.is_goal(node, node.route):
                continue

            # print("%d" % (len(goals)))

            for action in actions:

                # TODO: change to be from current to next planned step
                if step < len(route[0]):
                    planned_radians = math.atan2(route[0][step][0][1] - node.pos[0][1],
                                                 route[0][step][0][0] - node.pos[0][0])
                    planned_degrees = math.degrees(planned_radians)
                else:
                    planned_radians = math.atan2(route[0][-1][0][1] - node.pos[0][1],
                                                 route[0][-1][0][0] - node.pos[0][0])
                    planned_degrees = math.degrees(planned_radians)
                # print(planned_degrees)

                next_step_radians = math.atan2((node.pos[0][1] + action[1]) - node.pos[0][1],
                                               (node.pos[0][0] + action[0]) - node.pos[0][0])
                next_step_degrees = math.degrees(next_step_radians)

                anglediff = (next_step_degrees - planned_degrees + 180 + 360) % 360 - 180
                # print(anglediff, next_step_degrees, planned_degrees)

                if (anglediff > 90 or anglediff < -90):
                    continue

                #print('Here k = ', node.k, best_goal_value)
                if node.k == 0:

                    # no more err to use
                    pos = list()
                    for x in range(0, len(node.route)):
                        if step < len(node.route[x]):
                            pos.append(node.route[x][step][0])
                        else:
                            # TODO: Try to add none instead the last node and verify the code
                            pos.append(node.route[x][-1][0])
                    new_node = MDR.Node(pos, 0, step, node, node.route)
                    if (self.is_goal(new_node, new_node.route)):
                        goals.append(new_node)

                        # find the total cost of the current route
                        cost_new_node = MDR._compute_makespan(new_node.route) # TODO Change

                        if cost_new_node > best_goal_value:
                            best_goal_value = cost_new_node
                            #print('K= ', node.k, 'Best= ', best_goal_value)
                    else:
                        #f_score = self.f_interrupt(new_node, route) ## RONI: OLD LINE (bug: used route instead of new_node.route). #TODO: Refactor f_interrupt() so it accepts only the node
                        f_score = self.f_interrupt(new_node, new_node.route) ## RONI: NEW LINE
                        fscore[new_node] = f_score  # + node.step  # + (np.random.random()/1)
                        heapq.heappush(open, (fscore[new_node], new_node))

                    break

                # if action is legal (not collide in wall)
                if self.grid[node.pos[0][0] + action[0]][node.pos[0][1] + action[1]] == 1:
                    # array bound 1 on map
                    continue

                new_node, _ = self.apply(action, node, step, node.route)
                if new_node == None:
                    continue
                if (self.is_goal(new_node, new_node.route)):
                    goals.append(new_node)

                    # find the total cost of the current route
                    cost_new_node = MDR._compute_makespan(new_node.route)
                    #print('K= ', node.k, 'Best= ', best_goal_value)

                    if cost_new_node > best_goal_value:
                        best_goal_value = cost_new_node

                else:
                    f_score = self.f_interrupt(new_node, new_node.route)
                    fscore[new_node] = f_score # + node.step  # + (np.random.random()/1) # TODO: CHANGE 3
                    heapq.heappush(open, (fscore[new_node], new_node))

        # After all goals have been found
        max_step = 0
        best_goal = None
        for goal in goals:
            if goal.step > max_step:
                max_step = goal.step
                best_goal = goal

        # Reconstruct the solution
        solution = [best_goal]
        parent = best_goal.parent

        while parent is not None:
            solution.append(parent)
            parent = parent.parent

        solution = list(reversed(solution))

        return (solution, route, total_expanded_nodes)

    class Node:

        def __init__(self, pos, k, step, parent, route=None):
            self.pos = pos
            self.k = k
            self.step = step
            self.parent = parent
            self.route = route

        def weight(self):
            next_step_radians = math.atan2((self.route[0][-1][0][1]) - self.pos[0][1],
                                           (self.route[0][-1][0][0]) - self.pos[0][0])
            w = math.degrees(next_step_radians)

            return w

        def __lt__(self, other):
            return self.weight() < other.weight()



##############################################################################
# heuristic function for A* path scoring
##############################################################################
class Heuristic:
    def value(self, a, b): raise NotImplementedError

class ManhattanDistance(Heuristic):
    def value(self,a, b):
        distance = (abs(b[0] - a[0]) + abs(b[1] - a[1]))
        return distance

'''
    This heuristic accepts a database that maps a pair of nodes to the distance between them. 
'''
class DatabaseHeuristic(Heuristic):
    def __init__(self,goal_to_db):
        self.goal_to_db = goal_to_db

    def value(self,a, b):
        return self.goal_to_db[b][a]


##############################################################################
# path finding function
##############################################################################
def astar(array, start, goal, y, route, stp, h=ManhattanDistance):
    neighbors = [(0, 1), (0, -1), (1, 0), (-1, 0)]#, (0, 0)]#, (1, 1), (1, -1), (-1, 1), (-1, -1), (0, 0)]

    start_node = (start)

    close_set = set()
    came_from = {start_node: 0}
    gscore = {start_node: 0}
    fscore = {start_node: h.value(start[0], goal)}
    oheap = []
    heapq.heappush(oheap, (fscore[start_node], start_node))
    is_collision = 0

    while oheap:

        current = heapq.heappop(oheap)[1]

        if current[0] == goal:
            data = []
            while current in came_from:
                data.append(current)
                current = came_from[current]
            data = list(reversed(data))
            return data

        close_set.add(current)

        action_cost = 1 # TODO: Replace this in the future to support different action costs

        # Add stay at current position option
        neighbor = ((current[0][0], current[0][1]), current[1] + action_cost)
        came_from[neighbor] = (current)
        gscore[neighbor] = gscore[current] + action_cost
        fscore[neighbor] = gscore[neighbor] + h.value(neighbor[0], goal)
        heapq.heappush(oheap, (fscore[neighbor], neighbor))

        for i, j in neighbors:

            neighbor = ((current[0][0] + i, current[0][1] + j), current[1] + action_cost)
            tentative_g_score = gscore[current] + action_cost  # TODO-BUG: Here you will add square-root of 2 for a diagonal move. But, when you compute the g cost for the solution, you do not consider the cost of this move as square root of two but as one.


            ##Check for collision
            for a in range(0, y):#len(route)): # TODO-BUG: When you're re-planning for an agent, you need to check for all the other agents, not only those with a smaller index
                if len(route[a]) > neighbor[1]:
                    if neighbor[0] == route[a][neighbor[1]][0] and a != y: # TODO check the stp
                        is_collision = 1
                        break
            if is_collision == 1:
                is_collision = 0
                continue

            if 0 <= neighbor[0][0] < array.shape[0]:
                if 0 <= neighbor[0][1] < array.shape[1]:
                    if array[neighbor[0][0]][neighbor[0][1]] == 1:
                        # array bound 1 on map
                        continue
                else:
                    # array bound y walls
                    continue
            else:
                # array bound x walls
                continue

            if neighbor in close_set and tentative_g_score >= gscore.get(neighbor[0], 0):
                continue

            if tentative_g_score < gscore.get(neighbor[0], 0) or neighbor not in [i[1] for i in oheap]:
                came_from[neighbor] = (current)
                gscore[neighbor] = tentative_g_score
                fscore[neighbor] = tentative_g_score + h.value(neighbor[0], goal)
                heapq.heappush(oheap, (fscore[neighbor], neighbor))

    return False





'''
The entry point to running this algorithm
'''
def main():
    # Call a Workbook() function of openpyxl
    # to create a new blank Workbook object
    wb = openpyxl.Workbook()

    # Get workbook active sheet
    # from the active attribute.
    sheet = wb.active



    # load the map file into numpy array
    with open('Room1_new.txt', 'rt') as infile:
        grid1 = np.array([list(line.strip()) for line in infile.readlines()])
    print('Grid shape', grid1.shape)

    grid1[grid1 == '@'] = 1  # object on the map
    grid1[grid1 == 'T'] = 1  # object on the map
    grid1[grid1 == '.'] = 0  # free on map

    grid = np.array(grid1.astype(np.int))

    # Plot the map
    #fig, ax = plt.subplots(figsize=(10, 5))
    #ax.imshow(grid, cmap=plt.cm.ocean)

    # To convert new xlsx files into csv format, to use only once
    #path = 'room2\\'
    #extension = 'xlsx'
    #os.chdir(path)
    #excel_files = [i for i in glob.glob('*.{}'.format(extension))]
    #for excel in excel_files:
    #    out = excel.split('.')[0]+'.csv'
    #    df = pd.read_excel(excel) # if only the first sheet is needed.
    #    df.to_csv(out)

        # If you know the name of the column skip this
    #    first_column = df.columns[0]
    #    # Delete first
    #    df = df.drop([first_column], axis=1)
    #    df.to_csv(out, index=False)


    path = 'room1\\'
    extension = 'csv'
    os.chdir(path)
    files = [i for i in glob.glob('*.{}'.format(extension))]


    counter = 1

    res = []
    res = ['File Name', 'Leader No.', 'No. of Errors', 'No. of Agents', 'Len of Route', 'Run Time', 'Org. Cost', 'Total Expanded Nodes', 'No of Interrupt',
                       'New Score']
    # writing to the specified cell
    for len_of_row in range(1, 11):
        sheet.cell(row=counter, column=len_of_row).value = res[len_of_row-1]
    wb.save('results.xlsx')
    counter = counter+1


    for file in files:
        main_route = []
        new_route = []
        tmp_route = []
        #k = 3


        total_expanded_nodes = 0
        org_total_g = 0
        new_total_g = 0

        # reading csv file
        with open(file, 'rt') as csvfile:
            # creating a csv reader object
            csvreader = csv.reader(csvfile)
            header = next(csvreader)
            # extracting each data row one by one
            for row in csvreader:
                tmp_route = []
                for y in range(0, len(row)):
                    if row[y]:
                        tmp_route.append(ast.literal_eval(row[y]))
                main_route.append(tmp_route)

        org_total_g = max([len(route) for route in main_route])  # TODO Change

        ######LOOP Every agents as the leader######

        for current_agent in range(0, 1):#len(main_route)):

            temp_list_first = []
            temp_list_current = []
            temp_list_current = main_route[current_agent]
            temp_list_first = main_route[0]
            main_route[0] = temp_list_current
            main_route[current_agent] = temp_list_first


            ######LOOP for k between 1-5 errors######

            for k in range(1, 4):

                solution = []

                start_time = time.time()

                num_of_interrupts = 0
                print(file)
                print(counter - 1)

                mdr_solver = MDR()

                solution, route, total_expanded_nodes = mdr_solver.search_for_interrupt_plan(grid,main_route,k)

                elapsed_time = time.time() - start_time

                #for l in range(0, len(route)):
                #    print(route[l])

                print('Total runtime: '+ str(elapsed_time))
                total_routes_cost = sum([len(route) for route in solution[0].route])
                #print('Total routes cumulative sum: '+ str(total_routes_cost))
                print('Total expanded nodes: '+ str(total_expanded_nodes))
                print('Number of interruption for best solution: '+str(k - solution[-1].k))
                print('Number of Errors: ', k)
                new_total_g = 0


                for x in range(0, len(route)):
                    #print('**** Agent ', x, '****')
                    start = route[x][0][0]
                    goal = route[x][-1][0]
                    first_time_goal = 1
                    colors = ['r', 'purple', 'b', 'c', 'm', 'y', 'k', 'w']
                    ##############################################################################
                    # plot the path
                    ##############################################################################

                    # extract x and y coordinates from route list
                    x_coords = []
                    y_coords = []
                    for y in range(0, len(solution)):

                        x1 = solution[y].pos[x][0]
                        y1 = solution[y].pos[x][1]

                        if x1 == (route[x][-1][0][0]) and y1 == (route[x][-1][0][1]):
                            if first_time_goal == 0:
                                continue
                            else:
                                first_time_goal = 0
                        #else:
                        #    new_total_g = new_total_g + 1

                        x_coords.append(x1)
                        y_coords.append(y1)
                    new_total_g = max(new_total_g, len(x_coords)) # TODO CHnage

                print('Original g_score is - ', org_total_g)
                print('New g_score is - ', new_total_g)

                #for v in range(0, len(route)):
                #    print(route[v])


                no_of_interrupt = k - solution[-1].k
                res = [file, current_agent, k, len(route), len(route[1]), elapsed_time, org_total_g, total_expanded_nodes, no_of_interrupt,
                       new_total_g]

                # writing to the specified cell
                for len_of_row in range(1, 11):
                    sheet.cell(row=counter, column=len_of_row).value = res[len_of_row - 1]
                counter = counter + 1

            main_route[0] = temp_list_first
            main_route[current_agent] = temp_list_current

        wb.save('results.xlsx')




'''
Standard python code for making sure what should run when this module is run as the main script 
as oppose to when this code is imported. I'm adding this to allow unit testing.
'''
if __name__ == '__main__':
    main()